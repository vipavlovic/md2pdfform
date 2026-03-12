#!/usr/bin/env python3
"""
PDF Form to Excel Exporter (pdfform2excel.py)
Extracts filled form data from PDF files created by md2pdfform and exports to Excel
"""

import sys
import argparse
import os
import re
from pathlib import Path

try:
    from PyPDF2 import PdfReader
except ImportError:
    print("Error: PyPDF2 is required. Install with: pip install PyPDF2")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# --------------------------------------------------------------------------- #
# MD field extraction (used when --md is provided)
# --------------------------------------------------------------------------- #

def extract_field_names_from_md(md_path: str) -> list:
    """Return field names in the order they appear in the .md file."""
    text = Path(md_path).read_text(encoding="utf-8")

    entries = []  # (position, name)
    underscore_counter = 0

    for m in re.finditer(
        r'\{\{(?:text|email|number|date|textarea|checkbox|radio|dropdown)'
        r':([A-Za-z0-9_]+)',
        text,
    ):
        entries.append((m.start(), m.group(1)))

    for m in re.finditer(r'_{4,}', text):
        underscore_counter += 1
        entries.append((m.start(), f"field_{underscore_counter}"))

    entries.sort(key=lambda x: x[0])

    seen = set()
    ordered = []
    for _, name in entries:
        if name not in seen:
            seen.add(name)
            ordered.append(name)

    return ordered


# --------------------------------------------------------------------------- #
# Deduplication (used when --md is NOT provided)
# --------------------------------------------------------------------------- #

def deduplicate_fields(form_data: dict) -> dict:
    """
    Remove phantom fields introduced by two known md2pdfform artefacts:

    1. PyPDF2 radio-button duplicates – when acroForm.radio() is called once
       per option with the same field name, PyPDF2 renames duplicates as
       'name-1', 'name-2', etc.  The base 'name' already holds the correct
       selected value, so the suffixed copies are dropped.

    2. Fallback checkboxes – when acroForm.radio/choice fails, md2pdfform
       creates individual checkboxes named 'name_0', 'name_1', etc. (no base
       'name' exists).  These are merged back into a single 'name' entry whose
       value lists the checked option indices (or is empty when none checked).
    """
    result = {}

    # ---- pass 1: collect everything, skip PyPDF2 '-N' duplicates ----------
    # Pattern: ends with '-' followed by one or more digits
    dash_suffix = re.compile(r'^(.+)-(\d+)$')

    for name, value in form_data.items():
        m = dash_suffix.match(name)
        if m and m.group(1) in form_data:
            # Base name already present → this is a PyPDF2-renamed duplicate
            continue
        result[name] = value

    # ---- pass 2: merge '_N' fallback checkboxes into base name -------------
    # Pattern: ends with '_' followed by one or more digits
    under_suffix = re.compile(r'^(.+)_(\d+)$')

    groups = {}   # base_name → {index: value}
    lone = {}     # names that are NOT part of any '_N' group
    for name, value in result.items():
        m = under_suffix.match(name)
        if m:
            base = m.group(1)
            idx  = int(m.group(2))
            # Only treat as a group if the base name is NOT itself a real field
            if base not in result:
                groups.setdefault(base, {})[idx] = value
                continue
        lone[name] = value

    # Rebuild in original insertion order: lone fields first (preserving order),
    # then synthesised base names inserted at the position of their first member.
    merged = {}
    seen_bases = set()
    for name, value in result.items():
        m = under_suffix.match(name)
        if m:
            base = m.group(1)
            if base in groups and base not in seen_bases:
                seen_bases.add(base)
                # Value: the checked item(s), or empty if all unchecked
                checked = [
                    str(idx) for idx, v in sorted(groups[base].items())
                    if str(v).lower() not in ('', 'no', '/off', 'false', '0')
                ]
                merged[base] = ', '.join(checked)
            # Skip the individual _N entry
        elif name in lone:
            merged[name] = value

    return merged


# --------------------------------------------------------------------------- #
# Core extraction
# --------------------------------------------------------------------------- #

def sanitize_value(value):
    """Remove illegal characters for Excel cells"""
    if not isinstance(value, str):
        return value
    sanitized = ''.join(char for char in value if char.isprintable() or char in '\n\r\t')
    sanitized = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', sanitized)
    return sanitized


def extract_form_data(pdf_path, md_fields=None):
    """
    Extract form field data from a PDF file.

    If md_fields is provided (list of field names from the source .md file),
    only those fields are returned, in that order.  Any field present in the
    PDF but absent from md_fields is silently dropped.

    If md_fields is None, automatic deduplication is applied instead.
    """
    try:
        reader = PdfReader(pdf_path)

        if reader.is_encrypted:
            print(f"Warning: {pdf_path} is encrypted. Attempting to decrypt...")
            reader.decrypt('')

        if '/AcroForm' not in reader.trailer['/Root']:
            print(f"Warning: {pdf_path} does not contain form fields")
            return {}

        fields = reader.get_fields()

        if not fields:
            print(f"Warning: No form fields found in {pdf_path}")
            return {}

        # Build raw dict (preserving PDF order)
        raw = {}
        for field_name, field_info in fields.items():
            value = field_info.get('/V', '')

            if hasattr(value, 'get_object'):
                value = value.get_object()

            if isinstance(value, bytes):
                value = value.decode('utf-8', errors='ignore')

            if value == '/Yes':
                value = 'Yes'
            elif value == '/Off':
                value = 'No'

            raw[field_name] = sanitize_value(str(value) if value else '')

        if md_fields is not None:
            # Filter and reorder to match the MD file exactly
            md_set = set(md_fields)
            dropped = [n for n in raw if n not in md_set]
            if dropped:
                print(f"  Dropped {len(dropped)} unrecognised field(s): {', '.join(dropped)}")
            return {name: raw[name] for name in md_fields if name in raw}
        else:
            return deduplicate_fields(raw)

    except Exception as e:
        print(f"Error reading {pdf_path}: {e}")
        return {}


# --------------------------------------------------------------------------- #
# Export functions
# --------------------------------------------------------------------------- #

def export_single_pdf_to_excel(pdf_path, output_path, md_fields=None):
    """Export a single PDF form to Excel"""
    print(f"Processing: {pdf_path}")

    form_data = extract_form_data(pdf_path, md_fields)

    if not form_data:
        print("No form data found to export")
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form Data"

    ws['A1'] = 'Field Name'
    ws['B1'] = 'Value'

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['B1'].fill = header_fill
    ws['B1'].font = header_font

    for row, (field_name, value) in enumerate(form_data.items(), start=2):
        ws[f'A{row}'] = sanitize_value(field_name)
        ws[f'B{row}'] = sanitize_value(value)

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 60

    wb.save(output_path)
    print(f"Exported {len(form_data)} fields → {output_path}")
    return True


def export_multiple_pdfs_to_excel(pdf_paths, output_path, md_fields=None):
    """Export multiple PDF forms to a single Excel file with each PDF as a row"""
    print(f"Processing {len(pdf_paths)} PDF files...")

    all_field_names = {}  # ordered set
    all_data = []

    for pdf_path in pdf_paths:
        print(f"Reading: {pdf_path}")
        form_data = extract_form_data(pdf_path, md_fields)

        if form_data:
            all_field_names.update(dict.fromkeys(form_data.keys()))
            all_data.append({
                'filename': os.path.basename(pdf_path),
                'data': form_data
            })

    if not all_data:
        print("No form data found in any PDF")
        return False

    fields = list(all_field_names)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form Data"

    ws['A1'] = 'PDF Filename'
    for col_idx, field_name in enumerate(fields, start=2):
        ws.cell(row=1, column=col_idx, value=sanitize_value(field_name))

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(fields) + 2):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, entry in enumerate(all_data, start=2):
        ws.cell(row=row_idx, column=1, value=sanitize_value(entry['filename']))
        for col_idx, field_name in enumerate(fields, start=2):
            value = entry['data'].get(field_name, '')
            ws.cell(row=row_idx, column=col_idx, value=sanitize_value(value))

    ws.column_dimensions['A'].width = 30
    for col_idx in range(2, len(fields) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25

    wb.save(output_path)
    print(f"Exported {len(all_data)} PDFs × {len(fields)} fields → {output_path}")
    return True


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def main():
    parser = argparse.ArgumentParser(
        description='Extract filled PDF form data and export to Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Export single PDF
  %(prog)s filled_form.pdf -o output.xlsx

  # Export with source MD to filter/order fields correctly
  %(prog)s filled_form.pdf -o output.xlsx --md form.md

  # Export multiple PDFs to one Excel file (each PDF is a row)
  %(prog)s form1.pdf form2.pdf form3.pdf -o combined.xlsx --md form.md
        """
    )

    parser.add_argument('input', nargs='+', help='Input PDF file(s) with filled forms')
    parser.add_argument('-o', '--output', required=True, help='Output Excel file (.xlsx)')
    parser.add_argument(
        '--md',
        metavar='FILE',
        help='Source .md file used to generate the PDF form. '
             'When provided, only fields defined in the MD are exported '
             '(eliminates phantom fields and preserves MD field order).'
    )
    parser.add_argument(
        '--mode',
        choices=['single', 'combined'],
        default='auto',
        help='Export mode: single (one sheet per PDF) or combined (all PDFs in one sheet). Default: auto-detect'
    )

    args = parser.parse_args()

    # Validate inputs
    pdf_paths = []
    for path in args.input:
        if os.path.exists(path) and path.lower().endswith('.pdf'):
            pdf_paths.append(path)
        else:
            print(f"Warning: Skipping {path} (not found or not a PDF)")

    if not pdf_paths:
        print("Error: No valid PDF files provided")
        sys.exit(1)

    output_path = args.output
    if not output_path.lower().endswith('.xlsx'):
        output_path += '.xlsx'

    # Load MD field list if provided
    md_fields = None
    if args.md:
        if not Path(args.md).is_file():
            print(f"Error: MD file not found: {args.md}")
            sys.exit(1)
        md_fields = extract_field_names_from_md(args.md)
        print(f"Using MD field list ({len(md_fields)} fields) from: {args.md}")

    mode = args.mode
    if mode == 'auto':
        mode = 'single' if len(pdf_paths) == 1 else 'combined'

    try:
        if mode == 'single' and len(pdf_paths) == 1:
            success = export_single_pdf_to_excel(pdf_paths[0], output_path, md_fields)
        else:
            success = export_multiple_pdfs_to_excel(pdf_paths, output_path, md_fields)

        if success:
            print("\nExport completed successfully.")
        else:
            print("\nExport failed.")
            sys.exit(1)

    except Exception as e:
        print(f"\nError during export: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
