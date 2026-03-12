#!/usr/bin/env python3
"""
reorder_excel.py
Reorder rows/columns in a pdfform2excel output so fields follow the
same order as they appear in the source .md file.

Single-PDF layout  (col A = Field Name, col B = Value): reorders rows.
Multi-PDF layout   (row 1 = headers, each PDF is a row):  reorders columns.
"""

import sys
import argparse
import re
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# --------------------------------------------------------------------------- #
# MD field extraction
# --------------------------------------------------------------------------- #

def extract_field_names_from_md(md_path: str) -> list[str]:
    """Return field names in the order they appear in the .md file."""
    text = Path(md_path).read_text(encoding="utf-8")

    # Match all {{type:name:...}} fields and bare underscore sequences (____).
    # We collect (position, name) pairs then sort by position.
    entries: list[tuple[int, str]] = []
    underscore_counter = 0

    # Named fields: {{type:name[optional extra]}}
    for m in re.finditer(
        r'\{\{(?:text|email|number|date|textarea|checkbox|radio|dropdown)'
        r':([A-Za-z0-9_]+)',
        text,
    ):
        entries.append((m.start(), m.group(1)))

    # Underscore fields: 4 or more underscores → field_N
    for m in re.finditer(r'_{4,}', text):
        underscore_counter += 1
        entries.append((m.start(), f"field_{underscore_counter}"))

    # Sort by position, deduplicate while preserving first occurrence
    entries.sort(key=lambda x: x[0])
    seen: set[str] = set()
    ordered: list[str] = []
    for _, name in entries:
        if name not in seen:
            seen.add(name)
            ordered.append(name)

    return ordered


# --------------------------------------------------------------------------- #
# Excel layout detection
# --------------------------------------------------------------------------- #

def detect_layout(ws) -> str:
    """
    'single' – col A = 'Field Name', col B = 'Value'
    'multi'  – col A = 'PDF Filename', field names in B1, C1, …
    """
    a1 = str(ws["A1"].value or "").strip()
    if a1 == "Field Name":
        return "single"
    if a1 == "PDF Filename":
        return "multi"
    raise ValueError(
        f"Cannot detect layout: A1 contains '{a1}'. "
        "Expected 'Field Name' (single) or 'PDF Filename' (multi)."
    )


# --------------------------------------------------------------------------- #
# Reorder helpers
# --------------------------------------------------------------------------- #

def _copy_styles(src_cell, dst_cell):
    """Copy fill and font from src_cell to dst_cell."""
    if src_cell.has_style:
        dst_cell.font = src_cell.font.copy()
        dst_cell.fill = src_cell.fill.copy()
        dst_cell.border = src_cell.border.copy()
        dst_cell.alignment = src_cell.alignment.copy()


def reorder_single(ws, md_fields: list[str]):
    """Reorder rows in a single-PDF layout sheet."""
    # Read all data rows (skip header row 1)
    rows: dict[str, tuple] = {}  # field_name → (field_name_cell_value, value_cell_value)
    extra: list[tuple] = []       # rows whose field name is not in md_fields

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0] or "").strip()
        val  = row[1] if len(row) > 1 else ""
        if name in md_fields:
            rows[name] = (row[0], val)
        else:
            extra.append((row[0], val))

    # Build new ordering: md order first, then anything extra
    ordered_names = [n for n in md_fields if n in rows]
    new_rows = [rows[n] for n in ordered_names] + extra

    # Write back (starting at row 2)
    for i, (name_val, val) in enumerate(new_rows, start=2):
        ws.cell(row=i, column=1, value=name_val)
        ws.cell(row=i, column=2, value=val)


def reorder_multi(ws, md_fields: list[str]):
    """Reorder columns in a multi-PDF layout sheet (col A stays fixed)."""
    # Read current headers from row 1 (col B onwards)
    max_col = ws.max_column
    current_headers = []
    for col in range(2, max_col + 1):
        current_headers.append(str(ws.cell(row=1, column=col).value or "").strip())

    # Map header name → original column index (1-based)
    header_to_col: dict[str, int] = {h: col for col, h in enumerate(current_headers, start=2)}

    # Build new column order: md order first, then anything not in md_fields
    ordered_headers = [h for h in md_fields if h in header_to_col]
    extra_headers   = [h for h in current_headers if h not in set(md_fields)]
    new_headers     = ordered_headers + extra_headers

    if new_headers == current_headers:
        return  # nothing to do

    max_row = ws.max_row

    # Read all data for the columns we'll rearrange (col A is fixed)
    # data[col_idx] = list of (value, cell) for rows 1..max_row
    col_data: dict[int, list] = {}
    for col in range(2, max_col + 1):
        col_data[col] = [ws.cell(row=r, column=col).value for r in range(1, max_row + 1)]

    # Write columns back in new order
    for new_pos, header in enumerate(new_headers, start=2):
        old_col = header_to_col[header]
        src_values = col_data[old_col]
        for row_idx, val in enumerate(src_values, start=1):
            ws.cell(row=row_idx, column=new_pos, value=val)

    # Delete any now-surplus columns at the end (if new_headers is shorter — shouldn't happen)
    # (Covered implicitly because we overwrite in place using the same column count.)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main():
    parser = argparse.ArgumentParser(
        description="Reorder Excel form fields to match the source .md file order.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s form.md data.xlsx
  %(prog)s form.md data.xlsx -o reordered.xlsx
        """,
    )
    parser.add_argument("md_file",    help="Source .md file used to generate the PDF form")
    parser.add_argument("excel_file", help="Excel file produced by pdfform2excel.py")
    parser.add_argument(
        "-o", "--output",
        help="Output .xlsx path (default: overwrite input file)",
    )
    args = parser.parse_args()

    md_path    = args.md_file
    xlsx_path  = args.excel_file
    out_path   = args.output or xlsx_path

    # Validate inputs
    if not Path(md_path).is_file():
        print(f"Error: MD file not found: {md_path}")
        sys.exit(1)
    if not Path(xlsx_path).is_file():
        print(f"Error: Excel file not found: {xlsx_path}")
        sys.exit(1)

    print(f"Reading field order from: {md_path}")
    md_fields = extract_field_names_from_md(md_path)
    if not md_fields:
        print("Error: No form fields found in the MD file.")
        sys.exit(1)
    print(f"  Found {len(md_fields)} fields in MD")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    layout = detect_layout(ws)
    print(f"  Detected layout: {layout}")

    if layout == "single":
        reorder_single(ws, md_fields)
    else:
        reorder_multi(ws, md_fields)

    if not out_path.lower().endswith(".xlsx"):
        out_path += ".xlsx"

    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
